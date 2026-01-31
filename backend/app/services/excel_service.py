from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
from typing import List
from sqlalchemy.orm import Session


class ExcelService:
    @staticmethod
    def create_scan_report(
        db: Session, start_date: datetime = None, end_date: datetime = None
    ):
        """Create Excel report from scan records"""
        from ..models import ScanRecord

        query = db.query(ScanRecord)

        if start_date:
            query = query.filter(ScanRecord.scanned_at >= start_date)
        if end_date:
            query = query.filter(ScanRecord.scanned_at <= end_date)

        records = query.order_by(ScanRecord.scanned_at.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Report"

        # Headers
        headers = [
            "ID",
            "QR Content",
            "Scan Source",
            "User ID",
            "Scanned At",
            "Printed At",
            "Status",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.id)
            ws.cell(row=row, column=2, value=record.qr_content)
            ws.cell(row=row, column=3, value=record.scan_source)
            ws.cell(row=row, column=4, value=record.user_id)
            ws.cell(
                row=row,
                column=5,
                value=record.scanned_at.isoformat() if record.scanned_at else "",
            )
            ws.cell(
                row=row,
                column=6,
                value=record.printed_at.isoformat() if record.printed_at else "",
            )
            ws.cell(row=row, column=7, value=record.print_status)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output
